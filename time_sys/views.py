from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.contrib.auth.models import User
from django.utils import timezone
from datetime import datetime
import calendar

from store.models import Profile
from store.forms import UpdateUserForm
import pytz
# import json
#
from .models import TimeRecord, TimeTag, TimeReport, HealthMedicineTag, HealthRecord
from .forms import (TimeCheckoutForm, SettingsForm, DeleteTagForm, RecordDel,
                    MedicineForm, DeleteMedForm, MedicineRecordForm, WeightForm)
from .utils import (timezone_display, month_jump, generate_wheel, generate_wheel_with_rotation,
                    generate_monthly_report, get_or_create_time_report)


@login_required
# @user_passes_test(lambda u: u.is_superuser)
def time_checkout(request):
    form = TimeCheckoutForm(request.POST or None)
    form_del = RecordDel(request.POST or None)
    form_medicine = MedicineRecordForm(request.POST or None)
    form_weight = WeightForm(request.POST or None)


    user_id = request.user.id
    user_instance = User.objects.get(id=user_id)
    tags = TimeTag.objects.filter(user=user_instance)



    record_qs = TimeRecord.objects.filter(user=user_instance)
    checkout_time = timezone.now()

    this_utc = timezone.now()
    this_utc_ymd_str = this_utc.strftime('%Y-%m-%d')

    this_ym = timezone.now().strftime('%Y-%m')
    time_report_pr = get_or_create_time_report(user_instance, this_ym)
    day_data = time_report_pr.day_pr_data



    if this_utc_ymd_str not in day_data:
        day_data[this_utc_ymd_str] = []
        time_report_pr.day_pr_data = day_data
        time_report_pr.save()


    med_names = HealthMedicineTag.objects.filter(user=user_instance).all()


    # ----------------------------------#
    # New User Tags Creation #
    # ------------ ---- ---------------#
    if not tags.exists():
        messages.error(request, 'Please create a Time-Tag first.')
        return redirect('time_settings')

    # ----------------------------------#
    # Start Record #
    # ------------ ---- ---------------#
    if not record_qs.exists():
        TimeRecord.objects.create(
            user=user_instance,
            tag='start',
            end=checkout_time,  # <- this_utc, not datetime.this_utc()
            duration=0,
            type='UN'
        )
        messages.success(request, "Started first record.")
        return redirect('time_checkout')

    else:
        this_utc = timezone.now()
        last_utc = record_qs.order_by('-id').first().end

        if request.method == 'POST':
            form_type = request.POST.get("form_type")


            # ----------------------------------#
            # Checkout Form #
            # ------------ ---- ---------------#
            if form_type=="checkout" and form.is_valid():
                tag_raw = form.cleaned_data['time_tag'].split("-", 2)
                time_tag = tag_raw[0]
                time_type = tag_raw[1]

                # ----------------------------------#
                # Switch is on #
                # ------------ ----- ---------------#
                time_switch = form.cleaned_data['time_switch']
                if time_switch:
                    this_raw = form.cleaned_data['time_correction']
                    if timezone.is_aware(this_raw):
                        this_naive = this_raw.replace(tzinfo=None)
                        # // strip off the timezone info by using replace(tzinfo=None)
                    else:
                        this_naive = this_raw

                    user_tz = pytz.timezone(request.user.profile.timezone)
                    this_local = user_tz.localize(this_naive)

                    this_utc = this_local.astimezone(pytz.UTC)

                    if this_utc <= last_utc or this_utc > timezone.now():
                        messages.error(request, "Invalid time.")
                        return redirect('time_checkout')

                    duration = this_utc - last_utc
                    checkout_time = this_utc

                # ----------------------------------#
                # Switch is off #
                # ------------ Start ---------------#
                else:
                    duration = this_utc - last_utc
                    checkout_time = this_utc
                # END ------------------------------#


                this_utc_ymd_str = this_utc.strftime('%Y-%m-%d')
                last_utc_ymd_str = last_utc.strftime('%Y-%m-%d')


                # ----------------------------------#
                # Month Jump #
                # // by using "utlis.py"
                # ------------ ----- ---------------#
                month_jump_result = month_jump(last_utc, this_utc)
                if month_jump_result:
                    TimeRecord.objects.create(
                        user=user_instance,
                        tag=time_tag,
                        end=month_jump_result['end1'],
                        duration=month_jump_result['duration1'],
                        type=time_type
                    )
                    TimeRecord.objects.create(
                        user=user_instance,
                        tag=time_tag,
                        end=checkout_time,
                        duration=month_jump_result['duration2'],
                        type=time_type
                    )
                    first_checkout_dic = {this_utc_ymd_str:[]}
                    generate_monthly_report(k=user_instance, first_checkout_dic=first_checkout_dic)
                else:
                    TimeRecord.objects.create(
                        user=user_instance,
                        tag=time_tag,
                        end=checkout_time,
                        duration=int(duration.total_seconds()),
                        type=time_type
                    )

                # ----------------------------------#
                # PR in TimeReport
                # ------------ ---- ---------------#
                if time_type == "PR":
                    if last_utc_ymd_str not in day_data:
                        day_data[last_utc_ymd_str] = []
                    if this_utc_ymd_str != last_utc_ymd_str:
                        day_data[last_utc_ymd_str].append(time_tag)
                    else:
                        day_data[this_utc_ymd_str].append(time_tag)
                time_report_pr.day_pr_data = day_data
                time_report_pr.save()
                # END ------------------------------#

                return redirect('time_checkout')
            # END ------------------------------#


            # ----------------------------------#
            # Delete Form: del record by 'x' #
            # ----------------------------------#
            elif form_type=="del" and form_del.is_valid():
                record_del = TimeRecord.objects.filter(id=form_del.cleaned_data['del_id'], user=user_instance).first()
                if record_del.user.id == user_id:
                    record_del.delete()
                    return redirect('time_checkout')
            # END ------------------------------#


            # ----------------------------------#
            # Medicine Form #
            # ----------------------------------#
            elif form_type == "medicine" and form_medicine.is_valid():
                health_record, created = HealthRecord.objects.get_or_create(
                    day=this_utc_ymd_str,
                    user=user_instance,
                    defaults={'weight': {}, 'medicine': []}
                )
                # Safely get the current medicine data
                medicine_data = health_record.medicine or []
                # Update the data (e.g., mark that this medicine was taken today)
                medicine_data.append(form_medicine.cleaned_data['med_name'])  # Or any other value you want to store

                # Assign the updated dict back and save
                health_record.medicine = medicine_data
                health_record.save()
                return redirect('time_checkout')
            # END ------------------------------#


            # ----------------------------------#
            # Weight Form #
            # ----------------------------------#
            elif form_type == "weight" and form_weight.is_valid():
                # Get form data
                weight_morning_raw = form_weight.cleaned_data['weight_Morning']
                weight_beforelunch_raw = form_weight.cleaned_data['weight_BeforeLunch']
                weight_afterlunch_raw = form_weight.cleaned_data['weight_AfterLunch']
                weight_sleep_raw = form_weight.cleaned_data['weight_Sleep']

                # Helper function to parse and calculate fat weight
                # Helper function to parse and calculate fat weight
                def parse_weight_entry(entry):
                    if not entry:
                        return None  # Nothing entered
                    parts = entry.split()
                    try:
                        if len(parts) == 2:
                            weight = float(parts[0])
                            fat_percentage = float(parts[1])
                            fat_weight = round(weight * fat_percentage / 100, 2)
                            return [weight, fat_percentage, fat_weight]
                        elif len(parts) == 1:
                            weight = float(parts[0])
                            return [weight]
                        elif len(parts) == 3:
                            # Already split values, return as floats
                            return [float(parts[0]), float(parts[1]), float(parts[2])]
                        else:
                            return None  # Invalid format
                    except ValueError:
                        return None  # If not numeric input, ignore

                # Build the weight dict
                # Get or create
                health_record, created = HealthRecord.objects.get_or_create(
                    day=this_utc_ymd_str,
                    user=user_instance,
                    defaults={'weight': {}, 'medicine': []}
                )

                existing_weight = health_record.weight or {}

                # Initialize all keys, preserve previous data
                final_weight = {
                    "morning": existing_weight.get("morning", []),
                    "before lunch": existing_weight.get("before lunch", []),
                    "after lunch": existing_weight.get("after lunch", []),
                    "sleep": existing_weight.get("sleep", [])
                }

                # Update only submitted fields
                if weight_morning_raw:
                    final_weight["morning"] = parse_weight_entry(weight_morning_raw)
                    # // What you’re seeing is called the walrus operator (:=)
                    # // # Traditional way
                    # 1 result = parse_weight_entry(weight_morning_raw)
                    # 2 if result:
                    # 3     print(result)
                    # Walrus operator
                    # 1 if (result := parse_weight_entry(weight_morning_raw)):
                    # 2     print(result)
                if weight_beforelunch_raw:
                    final_weight["before lunch"] = parse_weight_entry(weight_beforelunch_raw)
                if weight_afterlunch_raw:
                    final_weight["after lunch"] = parse_weight_entry(weight_afterlunch_raw)
                if weight_sleep_raw:
                    final_weight["sleep"] = parse_weight_entry(weight_sleep_raw)

                # Save
                health_record.weight = final_weight
                health_record.save()

                return redirect('time_checkout')
            # END ------------------------------#


    if form.errors:
        for field, errs in form.errors.items():
            for err in errs:
                messages.error(request, f'{field}: {err}')

    # ----------------------------------#
    # 2 latest Time Records
    # ----------------------------------#
    records = TimeRecord.objects.filter(user=user_instance).order_by('-id')[:2]


    # ----------------------------------#
    # PR tags Polygon & Active
    # ----------------------------------#
    pr_tags = [i for i in tags if i.type == 'PR']
    #// special grammar
    n = len(pr_tags)
    length = 2000
    thickness1 = 80
    thickness2 = 80
    inner = 40
    wheel_fill = generate_wheel_with_rotation(n, length, thickness1+inner, thickness2+inner, pr_tags)
    angles = [(360 / n) * i for i in range(n)]
    # // special grammar

    tag_to_id = {tri['pr_tags'].tag.lower(): tri['id'] for tri in wheel_fill}
    #// to assign id number to each polygon button to control
    # END ------------------------------#


    # ----------------------------------#
    # Medicine & Active
    # ----------------------------------#
    health_record, created = HealthRecord.objects.get_or_create(
        day=this_utc_ymd_str,
        user=user_instance,
        defaults={'weight': {}, 'medicine': []}
    )
    today_meds = health_record.medicine
    # END ------------------------------#


    # ----------------------------------#
    # Weight Part 2: Pre-fill form data
    # ----------------------------------#
    health_record = HealthRecord.objects.filter(day=this_utc_ymd_str, user=user_instance).first()

    # Helper to convert the list to the correct input string
    def list_to_string(data):
        if not data:
            return ''
        # If only weight is present
        if len(data) == 1:
            return str(data[0])
        # If weight and fat % are present
        return ' '.join(str(i) for i in data)

    # Prepare initial data
    initial_weight_data = {}
    if health_record:
        weight = health_record.weight or {}
        initial_weight_data = {
            'weight_Morning': list_to_string(weight.get('morning')),
            'weight_BeforeLunch': list_to_string(weight.get('before lunch')),
            'weight_AfterLunch': list_to_string(weight.get('after lunch')),
            'weight_Sleep': list_to_string(weight.get('sleep')),
        }

    # Pass the initial data to the form
    form_weight = WeightForm(initial=initial_weight_data)
    # ----------------------------------#


    context = {
        'tag_to_id': tag_to_id,
        'wheel_fill': wheel_fill,
        'angles': angles,
        'wheel_ring': generate_wheel(n, length, thickness1, thickness2),
        'form': form,
        'form_del':form_del,
        'form_weight': form_weight,
        'weight_display': health_record.weight if health_record else {},
        'tags': tags,
        'timezone': timezone_display(request.user.profile),
        'now': this_utc,
        'last': last_utc if record_qs.exists() else None,
        'records': records,
        'today_tags': day_data[this_utc_ymd_str],
        'today_meds': list(set(today_meds)),
        'med_names': med_names,
    }
    return render(request, 'time/time_checkout.html', context)




@login_required
# @user_passes_test(lambda u: u.is_superuser)
def time_records(request):
    user_id = request.user.id
    user_instance = User.objects.get(id=user_id)
    start_of_day = timezone.now().replace(hour=0, minute=0, second=0, microsecond=0)
    records = TimeRecord.objects.filter(user=user_instance).order_by('-id')
    records_today = TimeRecord.objects.filter(user=user_instance, end__gte=start_of_day).order_by('-id')
    form = RecordDel(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        record_del = TimeRecord.objects.filter(id=form.cleaned_data['del_id'], user=user_instance).first()
        if record_del.user == request.user:
            record_del.delete()
        else:
            messages.success(request, "Not ur data")
    context = {
        'records': records,
        'records_today': records_today,
        'timezone': timezone_display(request.user.profile),
    }
    return render(request, 'time/time_records.html', context)





from django.http import JsonResponse
from django.utils.timezone import localtime

@login_required
def ajax_records_toggle(request):
    if request.method == "GET" and request.headers.get("x-requested-with") == "XMLHttpRequest":
        show_today = request.GET.get("today") == "1"
        user = request.user

        if show_today:
            today = timezone.localtime().date()
            records = TimeRecord.objects.filter(user=user, end__date=today).order_by('-end')
        else:
            records = TimeRecord.objects.filter(user=user).order_by('-end')

        user_timezone = pytz.timezone(user.profile.timezone)

        records_data = [{
            'id': r.id,
            'end': r.end.astimezone(user_timezone).strftime('%d | %H:%M:%S'),
            'duration': r.formatted_duration,
            'tag': r.tag,
            'type': r.type,
        } for r in records]

        return JsonResponse({'records': records_data})





import openpyxl
from django.http import HttpResponse
@login_required
def time_download(request):
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Time Records"

    # Query records
    # Determine which records to fetch
    if request.user.is_superuser:
        records = TimeRecord.objects.all().order_by('id')
    else:
        records = TimeRecord.objects.filter(user=request.user).order_by('-end')
    tz_str = request.user.profile.timezone
    user_tz = pytz.timezone(tz_str)

    # Add headers
    ws.append(["Id", "User", 'UTC', tz_str, "Duration", "Tag", "Type"])

    # Add rows
    for r in records:
        ws.append([
            r.id,
            r.user.id,
            localtime(r.end).strftime('%Y-%m-%d %H:%M:%S'),
            r.end.astimezone(user_tz).strftime('%Y-%m-%d %H:%M:%S'),
            r.formatted_duration,
            r.tag,
            r.type
        ])

    # Prepare response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=time_records.xlsx'
    wb.save(response)
    return response




from .forms import ExcelUploadForm
from .utils import parse_duration_to_seconds
@login_required
@user_passes_test(lambda u: u.is_superuser)
def upload_excel(request):
    if request.method == 'POST':
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            file = form.cleaned_data['file']
            wb = openpyxl.load_workbook(file)
            ws = wb.active

            rows = list(ws.iter_rows(min_row=2, values_only=True))  # skip header

            for row in rows:
                try:
                    record_id, user_id, utc_time_str, local_time_str, duration, tag, type_ = row

                    # Convert UTC time string to datetime
                    utc_time = timezone.datetime.strptime(utc_time_str, "%Y-%m-%d %H:%M:%S")
                    # utc_time = timezone.make_aware(utc_time, timezone.utc)

                    # Get user instance (assumes superuser or pre-existing users)
                    user = User.objects.get(id=user_id)

                    TimeRecord.objects.create(
                        user=user,
                        end=utc_time,
                        duration=parse_duration_to_seconds(duration),
                        tag=tag,
                        type=type_,
                    )
                except Exception as e:
                    messages.error(request, f"Error in row {row}: {e}")
                    continue

            messages.success(request, "Upload complete!")
            return redirect('time_records')
    else:
        form = ExcelUploadForm()

    return render(request, 'time/upload_excel.html', {'form': form})




@login_required
def time_report(request):
    user_instance = request.user
    this_utc = timezone.now()
    this_utc_ym_str = this_utc.strftime('%Y-%m')
    this_utc_ymd_str = this_utc.strftime('%Y-%m-%d')

    # Get the latest report
    reports_first = TimeReport.objects.filter(user=user_instance).order_by('-year_month').first()
    if reports_first and reports_first.year_month == this_utc_ym_str:
        generate_monthly_report(user_instance, reports_first.day_pr_data)

    # Prepare all reports
    reports = TimeReport.objects.filter(user=user_instance).order_by('-year_month')
    report_data = []
    for report in reports:
        total = report.total_duration or 1  # prevent division by zero
        tag_percent = {tag: [round(d / 60), round(d / total * 100, 2)] for tag, d in report.tag_data.items()}
        type_percent = {typ: [round(d / 60), round(d / total * 100, 2)] for typ, d in report.type_data.items()}
        report_data.append({'month': report.year_month, 'total_duration': report.total_duration, 'tag_percent': tag_percent, 'type_percent': type_percent})

    # Get PR tags
    pr_tags = TimeTag.objects.filter(user=user_instance, type='PR')

    # Get this month's health records
    health_records = HealthRecord.objects.filter(user=user_instance, day__startswith=this_utc_ym_str).order_by('-day')

    medicine_done = {}
    for health_record in health_records:
        medicine_done[health_record.day] = len(health_record.medicine)
    count_med_tags = len(HealthMedicineTag.objects.filter(user=user_instance))

    context = {
        'reports_first': reports_first,
        'pr_tags': pr_tags,
        'report_data': report_data,
        'timezone': timezone_display(user_instance.profile),
        'health_records': health_records,
        'medicine_done': medicine_done,
        'count_med_tags': count_med_tags
    }
    return render(request, 'time/time_report.html', context)






@login_required
def time_settings(request):
    form1 = SettingsForm(request.POST or None)
    form2 = DeleteTagForm(request.POST or None)
    form3 = UpdateUserForm(request.POST or None, instance=request.user.profile)
    form4 = MedicineForm(request.POST or None)
    del_med_form = DeleteMedForm(request.POST or None)
    user_id = request.user.id
    user_instance = User.objects.get(id=user_id)
    tags = TimeTag.objects.filter(user=user_instance)
    med_names = HealthMedicineTag.objects.filter(user=user_instance)

    if not tags.filter(user=user_instance).exists():
        tag_data = [
            {"user":user_instance, "tag": "Sleep", "type": "RS"},
            {"user":user_instance, "tag": "Eat", "type": "UN"},
            {"user":user_instance, "tag": "Transport", "type": "UN"},
            {"user":user_instance, "tag": "Nap", "type": "RS"},
            {"user":user_instance, "tag": "Shopping", "type": "UN"},
        ]

        for item in tag_data:
            TimeTag.objects.filter(user=user_instance).create(**item)

        tags = TimeTag.objects.filter(user=user_instance)

    if request.method == "POST":
        form_type = request.POST.get("form_type")

        if form_type == "settings_form":
            if form1.is_valid():
                tag_to_create = TimeTag(
                    user=user_instance,
                    tag=form1.cleaned_data['setting_tag'],
                    type=form1.cleaned_data['setting_type'],
                )
                tag_to_create.save()
                messages.success(request, f"You created tag: {tag_to_create.tag}")
                return redirect('time_settings')

        elif form_type == "delete_tag_form":
            if form2.is_valid():
                selected_tag = tags.filter(tag=form2.cleaned_data['selecting_tag'],
                                           user=user_instance).first()
                selected_tag.delete()
                messages.success(request, "Deleted")
                return redirect('time_settings')

        elif form_type == "change_timezone_form":
            if form3.is_valid():
                form3.save()
                messages.success(request, "Updated")
                return redirect('time_settings')

        elif form_type == "medicine_name_form":
            if form4.is_valid():
                tag_to_create = HealthMedicineTag(
                    user=user_instance,
                    name=form4.cleaned_data['medicine_name'],
                )
                tag_to_create.save()
                return redirect('time_settings')


        elif form_type == "delete_med_form":
            if del_med_form.is_valid():
                selected_tag = HealthMedicineTag.objects.filter(
                    id=del_med_form.cleaned_data['selecting_med'],
                    user=user_instance
                ).first()
                if selected_tag:
                    selected_tag.delete()
                    messages.success(request, "Deleted")
                else:
                    messages.error(request, "Tag not found.")
                return redirect('time_settings')

    context = {
        'form': form1,
        'form2': form2,
        'form3': form3,
        'form4': form4,
        'tags': tags,
        'med_names': med_names,
    }

    return render(request, 'time/time_settings.html', context)